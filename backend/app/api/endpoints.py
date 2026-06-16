from fastapi import APIRouter, Depends, Query, Form, HTTPException
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, desc, cast, Date
from datetime import datetime, timedelta
from typing import Optional
from fastapi.responses import FileResponse, JSONResponse
from passlib.hash import bcrypt
from jose import jwt, JWTError
from pydantic import BaseModel
import io, os, uuid, glob, tempfile, httpx

from ..database import get_db
from ..models import Sale, Order, Stock, DailySummary, SyncLog, UploadedReport, User
from ..services.wb_service import WBService
from ..config import settings

router = APIRouter(prefix="/api/v1")
security = HTTPBearer(auto_error=False)

JWT_SECRET = "wb-tools-secret-key-2024"
JWT_ALGORITHM = "HS256"
JWT_EXPIRE_HOURS = 24


class LoginRequest(BaseModel):
    username: str
    password: str


def create_token(username: str) -> str:
    expire = datetime.utcnow() + timedelta(hours=JWT_EXPIRE_HOURS)
    return jwt.encode({"sub": username, "exp": expire}, JWT_SECRET, algorithm=JWT_ALGORITHM)


async def get_current_user(credentials: Optional[HTTPAuthorizationCredentials] = Depends(security)) -> str:
    if credentials is None:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(credentials.credentials, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        return payload["sub"]
    except JWTError:
        raise HTTPException(status_code=401, detail="登录已过期，请重新登录")


@router.post("/auth/login")
async def login(req: LoginRequest, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(User).where(User.username == req.username))
    user = result.scalar_one_or_none()
    if not user or not bcrypt.verify(req.password, user.password_hash):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_token(user.username)
    return {"token": token, "username": user.username}


@router.get("/auth/verify")
async def verify(username: str = Depends(get_current_user)):
    return {"valid": True, "username": username}


# ── Sync ──────────────────────────────────────────────

@router.post("/sync/all")
async def sync_all(db: AsyncSession = Depends(get_db)):
    """Sync all data from WB API."""
    service = WBService(db, settings.wb_api_token)
    results = {}
    
    results["sales"] = await service.sync_sales()
    results["orders"] = await service.sync_orders()
    results["stocks"] = await service.sync_stocks()
    
    return results


@router.post("/sync/sales")
async def sync_sales(
    date_from: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
):
    service = WBService(db, settings.wb_api_token)
    return await service.sync_sales(date_from)


@router.post("/sync/orders")
async def sync_orders(
    date_from: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
):
    service = WBService(db, settings.wb_api_token)
    return await service.sync_orders(date_from)


@router.post("/sync/stocks")
async def sync_stocks(db: AsyncSession = Depends(get_db)):
    service = WBService(db, settings.wb_api_token)
    return await service.sync_stocks()


@router.post("/sync/summary")
async def sync_summary(date: str, db: AsyncSession = Depends(get_db)):
    service = WBService(db, settings.wb_api_token)
    return await service.calculate_daily_summary(date)


# ── Dashboard ─────────────────────────────────────────

@router.get("/dashboard")
async def get_dashboard(db: AsyncSession = Depends(get_db)):
    """Get dashboard key metrics."""
    today = datetime.utcnow().date()
    month_start = today.replace(day=1)
    
    # Today's stats from daily summaries
    today_summary = await db.execute(
        select(DailySummary).where(cast(DailySummary.date, Date) == today)
    )
    today_data = today_summary.scalar_one_or_none()
    
    # This month's stats
    month_summary = await db.execute(
        select(
            func.sum(DailySummary.total_sales).label("month_sales"),
            func.sum(DailySummary.total_for_pay).label("month_for_pay"),
            func.sum(DailySummary.total_orders).label("month_orders"),
            func.sum(DailySummary.total_commission).label("month_commission"),
            func.sum(DailySummary.total_cancelled).label("month_cancelled"),
        ).where(cast(DailySummary.date, Date) >= month_start)
    )
    month_data = month_summary.one()
    
    # Total stock value
    stock_query = await db.execute(
        select(func.sum(Stock.quantity * Stock.price).label("stock_value"),
               func.sum(Stock.quantity).label("total_quantity"))
    )
    stock_data = stock_query.one()
    
    # Recent sales trend (last 7 days)
    seven_days_ago = today - timedelta(days=7)
    trend_query = await db.execute(
        select(
            DailySummary.date,
            DailySummary.total_sales,
            DailySummary.total_for_pay,
            DailySummary.total_orders,
        ).where(
            cast(DailySummary.date, Date) >= seven_days_ago
        ).order_by(DailySummary.date)
    )
    trend = trend_query.all()
    
    return {
        "today": {
            "sales": round(today_data.total_sales, 2) if today_data else 0,
            "for_pay": round(today_data.total_for_pay, 2) if today_data else 0,
            "orders": today_data.total_orders if today_data else 0,
            "commission": round(today_data.total_commission, 2) if today_data else 0,
            "cancelled": today_data.total_cancelled if today_data else 0,
        } if today_data else {},
        "this_month": {
            "sales": round(month_data.month_sales or 0, 2),
            "for_pay": round(month_data.month_for_pay or 0, 2),
            "orders": int(month_data.month_orders or 0),
            "commission": round(month_data.month_commission or 0, 2),
            "cancelled": int(month_data.month_cancelled or 0),
        },
        "stocks": {
            "total_value": round(stock_data.stock_value or 0, 2),
            "total_quantity": int(stock_data.total_quantity or 0),
        },
        "trend": [
            {
                "date": r.date.strftime("%Y-%m-%d") if r.date else "",
                "sales": round(r.total_sales, 2),
                "for_pay": round(r.total_for_pay, 2),
                "orders": r.total_orders,
            }
            for r in trend
        ],
    }


# ── Sales ─────────────────────────────────────────────

@router.get("/sales")
async def list_sales(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    nm_id: Optional[int] = Query(None),
    db: AsyncSession = Depends(get_db),
):
    query = select(Sale).order_by(desc(Sale.date))
    
    if date_from:
        query = query.where(Sale.date >= datetime.strptime(date_from, "%Y-%m-%d"))
    if date_to:
        query = query.where(Sale.date <= datetime.strptime(date_to, "%Y-%m-%d") + timedelta(days=1))
    if nm_id:
        query = query.where(Sale.nm_id == nm_id)
    
    # Count
    count_query = select(func.count()).select_from(query.subquery())
    total = (await db.execute(count_query)).scalar()
    
    # Paginate
    query = query.offset((page - 1) * page_size).limit(page_size)
    result = await db.execute(query)
    sales = result.scalars().all()
    
    return {
        "items": [
            {
                "id": s.id,
                "sale_id": s.sale_id,
                "date": s.date.strftime("%Y-%m-%dT%H:%M:%S") if s.date else "",
                "supplier_article": s.supplier_article,
                "nm_id": s.nm_id,
                "barcode": s.barcode,
                "category": s.category,
                "subject": s.subject,
                "total_price": s.total_price,
                "discount_percent": s.discount_percent,
                "for_pay": s.for_pay,
                "finished_price": s.finished_price,
                "warehouse_name": s.warehouse_name,
            }
            for s in sales
        ],
        "total": total,
        "page": page,
        "page_size": page_size,
    }


# ── Orders ────────────────────────────────────────────

@router.get("/orders")
async def list_orders(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    is_cancel: Optional[bool] = Query(None),
    db: AsyncSession = Depends(get_db),
):
    query = select(Order).order_by(desc(Order.date))
    
    if date_from:
        query = query.where(Order.date >= datetime.strptime(date_from, "%Y-%m-%d"))
    if date_to:
        query = query.where(Order.date <= datetime.strptime(date_to, "%Y-%m-%d") + timedelta(days=1))
    if is_cancel is not None:
        query = query.where(Order.is_cancel == is_cancel)
    
    count_query = select(func.count()).select_from(query.subquery())
    total = (await db.execute(count_query)).scalar()
    
    query = query.offset((page - 1) * page_size).limit(page_size)
    result = await db.execute(query)
    orders = result.scalars().all()
    
    return {
        "items": [
            {
                "id": o.id,
                "date": o.date.strftime("%Y-%m-%dT%H:%M:%S") if o.date else "",
                "supplier_article": o.supplier_article,
                "nm_id": o.nm_id,
                "barcode": o.barcode,
                "total_price": o.total_price,
                "discount_percent": o.discount_percent,
                "finished_price": o.finished_price,
                "is_cancel": o.is_cancel,
                "warehouse_name": o.warehouse_name,
                "region_name": o.region_name,
            }
            for o in orders
        ],
        "total": total,
        "page": page,
        "page_size": page_size,
    }


# ── Stocks ────────────────────────────────────────────

@router.get("/stocks")
async def list_stocks(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    warehouse: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
):
    query = select(Stock).order_by(desc(Stock.quantity_full * Stock.price))
    
    if warehouse:
        query = query.where(Stock.warehouse_name == warehouse)
    
    count_query = select(func.count()).select_from(query.subquery())
    total = (await db.execute(count_query)).scalar()
    
    query = query.offset((page - 1) * page_size).limit(page_size)
    result = await db.execute(query)
    stocks = result.scalars().all()
    
    # Group by warehouse for summary
    warehouse_data = await db.execute(
        select(
            Stock.warehouse_name,
            func.sum(Stock.quantity * Stock.price).label("warehouse_value"),
            func.sum(Stock.quantity).label("warehouse_qty"),
        ).group_by(Stock.warehouse_name)
    )
    
    return {
        "items": [
            {
                "id": s.id,
                "nm_id": s.nm_id,
                "supplier_article": s.supplier_article,
                "barcode": s.barcode,
                "warehouse_name": s.warehouse_name,
                "quantity": s.quantity,
                "in_way_to_client": s.in_way_to_client,
                "in_way_from_client": s.in_way_from_client,
                "quantity_full": s.quantity_full,
                "price": s.price,
                "total_value": round((s.quantity or 0) * (s.price or 0), 2),
                "category": s.category,
                "subject": s.subject,
            }
            for s in stocks
        ],
        "total": total,
        "page": page,
        "page_size": page_size,
        "warehouses": [
            {"name": w.warehouse_name, "value": round(w.warehouse_value or 0, 2), "quantity": int(w.warehouse_qty or 0)}
            for w in warehouse_data.all()
        ],
    }


# ── Products Analysis ─────────────────────────────────

@router.get("/products/top")
async def get_top_products(
    days: int = Query(30, ge=1, le=365),
    limit: int = Query(10, ge=1, le=50),
    db: AsyncSession = Depends(get_db),
):
    """Get top products by sales."""
    since = datetime.utcnow() - timedelta(days=days)
    
    query = select(
        Sale.nm_id,
        Sale.supplier_article,
        Sale.subject,
        func.sum(Sale.total_price).label("total_revenue"),
        func.sum(Sale.for_pay).label("total_for_pay"),
        func.count(Sale.id).label("sale_count"),
        func.avg(Sale.discount_percent).label("avg_discount"),
    ).where(Sale.date >= since).group_by(
        Sale.nm_id, Sale.supplier_article, Sale.subject
    ).order_by(desc("total_revenue")).limit(limit)
    
    result = await db.execute(query)
    products = result.all()
    
    return {
        "items": [
            {
                "nm_id": p.nm_id,
                "article": p.supplier_article,
                "subject": p.subject,
                "revenue": round(p.total_revenue, 2),
                "for_pay": round(p.total_for_pay, 2),
                "sale_count": p.sale_count,
                "avg_discount": round(p.avg_discount, 1) if p.avg_discount else 0,
            }
            for p in products
        ]
    }


# ── Summary / Reports ─────────────────────────────────

@router.get("/reports/daily")
async def get_daily_reports(
    days: int = Query(30, ge=1, le=365),
    db: AsyncSession = Depends(get_db),
):
    since = datetime.utcnow().date() - timedelta(days=days)
    today = datetime.utcnow().date()
    
    query = select(DailySummary).where(
        cast(DailySummary.date, Date) >= since
    ).order_by(desc(DailySummary.date))
    
    result = await db.execute(query)
    reports = result.scalars().all()
    
    return {
        "items": [
            {
                "date": r.date.strftime("%Y-%m-%d") if r.date else "",
                "sales": round(r.total_sales, 2),
                "for_pay": round(r.total_for_pay, 2),
                "orders": r.total_orders,
                "commission": round(r.total_commission, 2),
                "cancelled": r.total_cancelled,
                "cancelled_amount": round(r.total_cancelled_amount, 2),
                "avg_discount": round(r.avg_discount, 1),
                "products": r.product_count,
            }
            for r in reports
        ]
    }


@router.get("/reports/summary")
async def get_summary_report(
    days: int = Query(30, ge=1, le=365),
    db: AsyncSession = Depends(get_db),
):
    since_date = (datetime.utcnow() - timedelta(days=days)).date()
    
    query = select(
        func.sum(DailySummary.total_sales).label("total_sales"),
        func.sum(DailySummary.total_for_pay).label("total_for_pay"),
        func.sum(DailySummary.total_commission).label("total_commission"),
        func.sum(DailySummary.total_orders).label("total_orders"),
        func.sum(DailySummary.total_cancelled).label("total_cancelled"),
        func.avg(DailySummary.avg_discount).label("avg_discount"),
    ).where(cast(DailySummary.date, Date) >= since_date)
    
    result = await db.execute(query)
    r = result.one()
    
    # Count active products
    product_count_q = await db.execute(
        select(func.count(func.distinct(Sale.nm_id)))
        .where(cast(Sale.date, Date) >= since_date)
    )
    product_count = product_count_q.scalar()
    
    return {
        "period_days": days,
        "total_sales": round(r.total_sales or 0, 2),
        "net_revenue": round(r.total_for_pay or 0, 2),
        "total_commission": round(r.total_commission or 0, 2),
        "total_orders": int(r.total_orders or 0),
        "total_cancelled": int(r.total_cancelled or 0),
        "cancel_rate": round((r.total_cancelled or 0) / max(r.total_orders or 1, 1) * 100, 1),
        "avg_discount": round(r.avg_discount or 0, 1),
        "active_products": product_count or 0,
    }


# ── Sync Logs ─────────────────────────────────────────

@router.get("/sync/logs")
async def get_sync_logs(
    limit: int = Query(20, ge=1, le=100),
    db: AsyncSession = Depends(get_db),
):
    query = select(SyncLog).order_by(desc(SyncLog.id)).limit(limit)
    result = await db.execute(query)
    logs = result.scalars().all()
    
    return {
        "items": [
            {
                "id": l.id,
                "data_type": l.data_type,
                "records_count": l.records_count,
                "status": l.status,
                "error_message": l.error_message,
                "created_at": l.created_at.strftime("%Y-%m-%dT%H:%M:%S") if l.created_at else "",
            }
            for l in logs
        ]
    }


# ── Warehouse list ────────────────────────────────────

@router.get("/warehouses")
async def get_warehouses(db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(Stock.warehouse_name, func.count(Stock.id).label("count"))
        .group_by(Stock.warehouse_name)
        .order_by(Stock.warehouse_name)
    )
    return {"items": [{"name": r.warehouse_name, "count": r.count} for r in result.all()]}


# ── Report Aggregation (Weekly/Monthly) ────────────────

@router.get("/reports/aggregate")
async def get_aggregated_reports(
    year: int = Query(...),
    month: int = Query(...),
    db: AsyncSession = Depends(get_db),
):
    """Get weekly and monthly aggregated reports from API data + uploaded reports."""
    from sqlalchemy import extract
    
    # Date range for the month
    month_start = datetime(year, month, 1)
    if month == 12:
        month_end = datetime(year + 1, 1, 1)
    else:
        month_end = datetime(year, month + 1, 1)
    
    # Get daily summaries for this month
    daily_q = await db.execute(
        select(DailySummary).where(
            DailySummary.date >= month_start,
            DailySummary.date < month_end,
        ).order_by(DailySummary.date)
    )
    daily = daily_q.scalars().all()
    
    # Get uploaded reports for this month
    upload_q = await db.execute(
        select(UploadedReport).where(
            UploadedReport.period_start >= month_start,
            UploadedReport.period_start < month_end,
        ).order_by(UploadedReport.period_start)
    )
    uploads = upload_q.scalars().all()
    
    # Build weekly groups
    weeks = {}
    for d in daily:
        week_num = d.date.isocalendar()[1]
        if week_num not in weeks:
            weeks[week_num] = {
                "week_num": week_num,
                "start": None,
                "end": None,
                "sales": 0, "for_pay": 0, "commission": 0, "orders": 0, "cancelled": 0,
            }
        w = weeks[week_num]
        if w["start"] is None or d.date < w["start"]:
            w["start"] = d.date
        if w["end"] is None or d.date > w["end"]:
            w["end"] = d.date
        w["sales"] += d.total_sales or 0
        w["for_pay"] += d.total_for_pay or 0
        w["commission"] += d.total_commission or 0
        w["orders"] += d.total_orders or 0
        w["cancelled"] += d.total_cancelled or 0
    
    # Monthly totals
    month_sales = sum(d.total_sales or 0 for d in daily)
    month_for_pay = sum(d.total_for_pay or 0 for d in daily)
    month_commission = sum(d.total_commission or 0 for d in daily)
    month_orders = sum(d.total_orders or 0 for d in daily)
    month_cancelled = sum(d.total_cancelled or 0 for d in daily)
    
    # Uploaded report totals
    upload_sales = sum(u.total_sales or 0 for u in uploads)
    upload_for_pay = sum(u.total_for_pay or 0 for u in uploads)
    
    def fmt_date(dt):
        return dt.strftime("%m-%d") if dt else ""
    
    return {
        "year": year,
        "month": month,
        "weeks": [
            {
                "label": f"第{w['week_num']}周 ({fmt_date(w['start'])}~{fmt_date(w['end'])})",
                "sales": round(w["sales"], 2),
                "for_pay": round(w["for_pay"], 2),
                "commission": round(w["commission"], 2),
                "orders": w["orders"],
                "cancelled": w["cancelled"],
            }
            for w in sorted(weeks.values(), key=lambda x: x["week_num"])
        ],
        "monthly": {
            "sales": round(month_sales, 2),
            "for_pay": round(month_for_pay, 2),
            "commission": round(month_commission, 2),
            "orders": month_orders,
            "cancelled": month_cancelled,
            "from_api": True,
        },
        "uploaded": [
            {
                "id": u.id,
                "filename": u.filename,
                "period": f"{u.period_start.strftime('%m-%d')} ~ {u.period_end.strftime('%m-%d')}" if u.period_start else "",
                "sales": round(u.total_sales, 2),
                "for_pay": round(u.total_for_pay, 2),
            }
            for u in uploads
        ],
        "upload_totals": {
            "sales": round(upload_sales, 2),
            "for_pay": round(upload_for_pay, 2),
        },
    }


# ── Excel Upload ───────────────────────────────────────

import os
import uuid
from fastapi import UploadFile, File, HTTPException, Form
from openpyxl import load_workbook

UPLOAD_DIR = "/usr/share/nginx/html/uploads/reports/"

# WB API JSON → Excel column mapping for reportDetailByPeriod
WB_REPORT_FIELD_MAP = {
    "№": lambda r: str(r.get("gi_id", "")),
    "номер поставки": lambda r: str(r.get("gi_box_type_name", "")),
    "предмет": lambda r: r.get("subject", ""),
    "код номенклатуры": lambda r: str(r.get("nm_id", "")),
    "бренд": lambda r: r.get("brand_name", ""),
    "артикул поставщика": lambda r: r.get("sa_name", ""),
    "название": lambda r: r.get("ts_name", ""),
    "размер": lambda r: r.get("ts_name", ""),
    "баркод": lambda r: str(r.get("barcode", "")),
    "тип документа": lambda r: r.get("doc_type_name", ""),
    "обоснование для оплаты": lambda r: r.get("operation_type", ""),
    "дата заказа покупателем": lambda r: (r.get("order_dt") or "")[:10] if r.get("order_dt") else "",
    "дата продажи": lambda r: (r.get("sale_dt") or "")[:10] if r.get("sale_dt") else "",
    "кол-во": lambda r: r.get("quantity", 0),
    "цена розничная": lambda r: r.get("retail_price", 0),
    "вайлдберриз реализовал товар (пр)": lambda r: r.get("retail_amount", 0),
    "к перечислению продавцу за реализованный товар": lambda r: r.get("ppvz_for_pay", 0),
    "услуги по доставке товара покупателю": lambda r: r.get("delivery_rub", 0),
    "количество доставок": lambda r: r.get("delivery_amount", 0),
    "количество возврата": lambda r: r.get("return_amount", 0),
    "вознаграждение с продаж до вычета услуг поверенного, без ндс": lambda r: r.get("ppvz_vw", 0),
    "хранение": lambda r: r.get("storage_fee", 0),
    "штрафы": lambda r: r.get("penalty", 0),
    "удержания": lambda r: r.get("additional_payment", 0),
    "возмещение за выдачу и возврат товаров на пвз": lambda r: r.get("rebill_logistic_cost", 0),
}


@router.post("/reports/fetch-auto")
async def fetch_and_process(
    start_date: str = Form(...),
    end_date: str = Form(...),
    exchange_rate: float = Form(12.5),
    tax_rate: float = Form(12),
    fee_rate: float = Form(7),
    db: AsyncSession = Depends(get_db),
):
    """Fetch WB report for date range via API, convert to Excel, process into profit table."""
    service = WBService(db, settings.wb_api_token)
    
    try:
        raw_data = await service.client.get_report_detail(start_date, end_date)
    except httpx.HTTPStatusError as e:
        if e.response.status_code == 429:
            raise HTTPException(status_code=429, detail=f"WB API 请求过于频繁（429），请稍等 1-2 分钟后重试")
        raise HTTPException(status_code=502, detail=f"WB API 请求失败: {str(e)}")
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"WB API 请求失败: {str(e)}")
    
    if not raw_data:
        raise HTTPException(status_code=404, detail=f"{start_date} ~ {end_date} 期间没有报表数据")
    
    # Build column headers (Russian order matching WB report export)
    from openpyxl import Workbook
    russian_headers = list(WB_REPORT_FIELD_MAP.keys())
    
    wb_in = Workbook()
    ws_in = wb_in.active
    ws_in.title = "Report"
    ws_in.append(russian_headers)
    
    for record in raw_data:
        row = []
        for h in russian_headers:
            try:
                val = WB_REPORT_FIELD_MAP[h](record)
            except Exception:
                val = ""
            row.append(val)
        ws_in.append(row)
    
    # Save the generated Excel file
    os.makedirs(UPLOAD_DIR, exist_ok=True)
    source_fname = f"WB_API_{start_date}_{end_date}.xlsx"
    source_path = os.path.join(UPLOAD_DIR, source_fname)
    wb_in.save(source_path)
    
    # Save as UploadedReport (platform type)
    report = UploadedReport(
        filename=source_fname,
        file_type="platform",
        period_start=datetime.strptime(start_date, "%Y-%m-%d"),
        period_end=datetime.strptime(end_date, "%Y-%m-%d"),
        total_sales=0,
        total_for_pay=0,
        records_count=len(raw_data),
        source="api",
        file_path=source_path,
    )
    db.add(report)
    await db.commit()
    await db.refresh(report)
    
    # Use shared processing function (same logic as manual upload)
    purchase_data = await _load_purchase_data(db)
    result = await _process_single_report(
        source_path, purchase_data,
        exchange_rate, tax_rate, fee_rate, db,
        cleanup_ids=[report.id],
        cleanup_paths=[source_path],
    )
    result["stats"]["start_date"] = start_date
    result["stats"]["end_date"] = end_date
    result["stats"]["source_records"] = len(raw_data)
    return result


@router.post("/reports/upload")
async def upload_report(
    file: UploadFile = File(...),
    file_type: str = Form("platform"),  # "platform" or "purchase"
    db: AsyncSession = Depends(get_db),
):
    """Upload a WB report Excel file (platform) or purchase file (cost+head freight)."""
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx / .xls 文件")
    
    os.makedirs(UPLOAD_DIR, exist_ok=True)
    file_path = os.path.join(UPLOAD_DIR, f"{uuid.uuid4()}_{file.filename}")
    
    content = await file.read()
    with open(file_path, "wb") as f:
        f.write(content)
    
    try:
        # For purchase files, just save path, minimal parsing
        if file_type == "purchase":
            # Parse purchase file to extract product cost data
            purchase_data = []
            wb = load_workbook(file_path, data_only=True)
            ws = wb.active
            headers = []
            for row in ws.iter_rows(values_only=True):
                if not row or not any(row):
                    continue
                if not headers:
                    headers = [str(c or "").strip() for c in row]
                    continue
                purchase_data.append({headers[i]: str(row[i] or "") if i < len(row) else "" for i in range(len(headers))})
            
            report = UploadedReport(
                filename=file.filename,
                file_type="purchase",
                period_start=datetime.utcnow(),
                period_end=datetime.utcnow(),
                total_sales=0,
                total_for_pay=0,
                records_count=len(purchase_data),
                source="upload",
                file_path=file_path,
            )
            db.add(report)
            await db.commit()
            await db.refresh(report)
            
            return {
                "status": "success",
                "id": report.id,
                "filename": file.filename,
                "file_type": "purchase",
                "rows": len(purchase_data),
            }
        
        # Platform file parsing
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
        
        total_sales = 0.0
        total_for_pay = 0.0
        total_commission = 0.0
        total_logistics = 0.0
        total_storage = 0.0
        row_count = 0
        period_start = None
        period_end = None
        
        col_map = {}
        KNOWN_COLUMNS = {
            "вайлдберриз реализовал товар (пр)": "sales",
            "к перечислению продавцу за реализованный товар": "for_pay",
            "вознаграждение с продаж до вычета услуг поверенного, без ндс": "commission",
            "услуги по доставке товара покупателю": "logistics",
            "хранение": "storage",
            "возмещение за выдачу и возврат товаров на пвз": "returns_fee",
            "дата продажи": "sale_date",
            "дата заказа покупателем": "order_date",
            "тип документа": "doc_type",
            "кол-во": "quantity",
            "цена розничная": "retail_price",
        }
        header_found = False
        
        for row in ws.iter_rows(values_only=True):
            if not row or not any(row):
                continue
            
            row_count += 1
            
            if not header_found:
                for i, cell in enumerate(row):
                    if cell and isinstance(cell, str):
                        key = cell.strip().lower()
                        if key in KNOWN_COLUMNS:
                            col_map[KNOWN_COLUMNS[key]] = i
                            header_found = True
                if header_found:
                    continue
            
            if not header_found:
                continue
            
            try:
                if "sales" in col_map:
                    v = row[col_map["sales"]]
                    if isinstance(v, (int, float)):
                        total_sales += float(v)
                if "for_pay" in col_map:
                    v = row[col_map["for_pay"]]
                    if isinstance(v, (int, float)):
                        total_for_pay += float(v)
                if "commission" in col_map:
                    v = row[col_map["commission"]]
                    if isinstance(v, (int, float)):
                        total_commission += float(v)
                if "logistics" in col_map:
                    v = row[col_map["logistics"]]
                    if isinstance(v, (int, float)):
                        total_logistics += float(v)
                if "storage" in col_map:
                    v = row[col_map["storage"]]
                    if isinstance(v, (int, float)):
                        total_storage += float(v)
            except (ValueError, TypeError, IndexError):
                pass
        
        report = UploadedReport(
            filename=file.filename,
            file_type="platform",
            period_start=period_start or datetime.utcnow(),
            period_end=period_end or datetime.utcnow(),
            total_sales=total_sales,
            total_for_pay=total_for_pay,
            total_commission=total_commission,
            total_logistics=total_logistics,
            total_storage=total_storage,
            records_count=row_count,
            source="upload",
            file_path=file_path,
        )
        db.add(report)
        await db.commit()
        await db.refresh(report)
        
        return {
            "status": "success",
            "id": report.id,
            "filename": file.filename,
            "file_type": "platform",
            "total_sales": round(total_sales, 2),
            "total_for_pay": round(total_for_pay, 2),
            "rows": row_count,
        }
    
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"解析失败: {str(e)}")


@router.post("/reports/uploaded/{report_id}/delete")
async def delete_uploaded_report(
    report_id: int,
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(select(UploadedReport).where(UploadedReport.id == report_id))
    report = result.scalar_one_or_none()
    if not report:
        raise HTTPException(status_code=404, detail="报表不存在")
    await db.delete(report)
    await db.commit()
    return {"status": "success", "id": report_id}


@router.get("/reports/uploaded")
async def list_uploaded_reports(
    limit: int = Query(20, ge=1, le=100),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(
        select(UploadedReport).order_by(desc(UploadedReport.uploaded_at)).limit(limit)
    )
    reports = result.scalars().all()
    
    return {
        "items": [
            {
                "id": r.id,
                "filename": r.filename,
                "period_start": r.period_start.strftime("%Y-%m-%d") if r.period_start else "",
                "period_end": r.period_end.strftime("%Y-%m-%d") if r.period_end else "",
                "total_sales": round(r.total_sales, 2),
                "total_for_pay": round(r.total_for_pay, 2),
                "total_commission": round(r.total_commission, 2),
                "total_logistics": round(r.total_logistics, 2),
                "total_storage": round(r.total_storage, 2),
                "records_count": r.records_count,
                "source": r.source,
                "uploaded_at": r.uploaded_at.strftime("%m-%d %H:%M") if r.uploaded_at else "",
            }
            for r in reports
        ]
    }


# ── Product Analysis from Uploaded Excel ──────────────────

import glob as glob_module
import os as os_module

@router.get("/reports/products")
async def get_product_analysis(
    db: AsyncSession = Depends(get_db),
):
    """Parse the latest uploaded Excel file and return product-level aggregation."""
    upload_dir = UPLOAD_DIR
    files = glob_module.glob(os_module.path.join(upload_dir, "*.xlsx"))
    if not files:
        raise HTTPException(status_code=404, detail="没有找到已上传的 Excel 文件")
    
    latest = max(files, key=os_module.path.getmtime)
    
    try:
        wb = load_workbook(latest, data_only=True)
        ws = wb.active
        
        col_map = {}
        KNOWN = {
            "артикул поставщика": "article",
            "название": "name",
            "предмет": "category",
            "размер": "size",
            "бренд": "brand",
            "вайлдберриз реализовал товар (пр)": "sales",
            "к перечислению продавцу за реализованный товар": "for_pay",
            "кол-во": "qty",
            "цена розничная": "retail_price",
            "вознаграждение с продаж до вычета услуг поверенного, без ндс": "commission",
            "услуги по доставке товара покупателю": "logistics",
        }
        
        header_found = False
        products = {}
        total_sales = 0.0
        total_for_pay = 0.0
        
        for row in ws.iter_rows(values_only=True):
            if not row or not any(row):
                continue
            
            if not header_found:
                for i, cell in enumerate(row):
                    if cell and isinstance(cell, str):
                        key = cell.strip().lower()
                        if key in KNOWN:
                            col_map[KNOWN[key]] = i
                            header_found = True
                if header_found:
                    continue
                continue
            
            try:
                article = str(row[col_map.get("article", -1)] or "").strip() if "article" in col_map else ""
                name = str(row[col_map.get("name", -1)] or "").strip() if "name" in col_map else ""
                category = str(row[col_map.get("category", -1)] or "").strip() if "category" in col_map else ""
                
                if not article:
                    continue
                
                sales = float(row[col_map["sales"]]) if "sales" in col_map and isinstance(row[col_map["sales"]], (int, float)) else 0
                for_pay = float(row[col_map["for_pay"]]) if "for_pay" in col_map and isinstance(row[col_map["for_pay"]], (int, float)) else 0
                qty = float(row[col_map.get("qty", -1)] or 0) if "qty" in col_map and isinstance(row[col_map.get("qty", -1)], (int, float)) else 0
                
                if article not in products:
                    products[article] = {
                        "article": article,
                        "name": name or article,
                        "category": category,
                        "sales": 0.0,
                        "for_pay": 0.0,
                        "qty": 0,
                        "variants": 0,
                        "sizes": set(),
                    }
                
                p = products[article]
                p["sales"] += sales
                p["for_pay"] += for_pay
                p["qty"] += qty
                total_sales += sales
                total_for_pay += for_pay
                
                if "size" in col_map:
                    sz = str(row[col_map["size"]] or "").strip()
                    if sz:
                        p["sizes"].add(sz)
                
            except (ValueError, TypeError, IndexError):
                pass
        
        result = []
        for article, p in products.items():
            p["variants"] = len(p["sizes"])
            del p["sizes"]
            result.append(p)
        
        result.sort(key=lambda x: x["sales"], reverse=True)
        
        return {
            "products": result,
            "total_products": len(result),
            "total_sales": round(total_sales, 2),
            "total_for_pay": round(total_for_pay, 2),
        }
    
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"解析失败: {str(e)}")


# ── Multi-sheet Processing & Download ────────────────────
# Translates raw WB report → 初处理 → 销售/物流/仓储/广告 → 利润表

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import re

# Russian → Chinese header mapping
HEADER_MAP = {
    "№": "报告编号",
    "номер поставки": "供货编号",
    "предмет": "商品品类",
    "код номенклатуры": "wb商品编号",
    "бренд": "品牌",
    "артикул поставщика": "卖家商品编号",
    "название": "商品名称",
    "размер": "尺寸",
    "баркод": "条形码",
    "тип документа": "文档类型",
    "обоснование для оплаты": "付款依据",
    "дата заказа покупателем": "买家下单日期",
    "дата продажи": "销售日期",
    "кол-во": "数量",
    "цена розничная": "零售价",
    "вайлдберриз реализовал товар (пр)": "wb销售商品",
    "согласованный продуктовый дисконт, %": "协商商品折扣%",
    "промокод, %": "促销码%",
    "итоговая согласованная скидка, %": "最终协商折扣%",
    "цена розничная с учетом согласованной скидки": "零售价含协商折扣",
    "размер снижения квв из-за рейтинга, %": "评级减少佣金%",
    "размер изменения квв из-за акции, %": "促销减少佣金%",
    "платформенные скидки, %": "常客折扣%",
    "размер квв, %": "wb佣金比例%",
    "размер квв без ндс, % базовый": "wb基础佣金(无增值税)%",
    "итоговый квв без ндс, %": "wb最终佣金(无增值税)%",
    "вознаграждение с продаж до вычета услуг поверенного, без ндс": "未扣服务费的销售佣金",
    "возмещение за выдачу и возврат товаров на пвз": "取件退货赔偿",
    "компенсация платёжных услуг/комиссия за интеграцию платёжных сервисов": "收单支付管理费",
    "размер компенсации платёжных услуг/комиссии за интеграцию платёжных сервисов, %": "收单服务费比例%",
    "тип платежа: компенсация платёжных услуг/комиссия за интеграцию платёжных сервисов": "收单支付类型",
    "вознаграждение вайлдберриз (вв), без ндс": "wb佣金(无增值税)",
    "ндс с вознаграждения вайлдберриз": "wb佣金增值税",
    "к перечислению продавцу за реализованный товар": "支付给卖家的已售商品金额",
    "количество доставок": "交付数量",
    "количество возврата": "退货数量",
    "услуги по доставке товара покупателю": "向买家交付货物的服务",
    "дата начала действия фиксации": "承诺开始日期",
    "дата конца действия фиксации": "固定折扣截止日期",
    "признак услуги платной доставки": "付费送货标志",
    "общая сумма штрафов": "罚款总额",
    "корректировка вознаграждения вайлдберриз (вв)": "wb佣金调整",
    "виды логистики, штрафов и корректировок вв": "wb物流罚款调整类型",
    "стикер мп": "fbs标签",
    "наименование банка-эквайера": "收单银行名称",
    "номер офиса": "办公室号码",
    "наименование офиса доставки": "送货点名称",
    "инн партнера": "合作站点税号",
    "партнер": "合作站点",
    "склад": "仓库",
    "страна": "国家",
    "тип коробов": "包装盒类型",
    "номер таможенной декларации": "报关单号",
    "номер сборочного задания": "装配任务编号",
    "код маркировки": "标记编码",
    "шк": "仓库标签",
    "srid": "srid",
    "возмещение издержек по перевозке/по складским операциям с товаром": "运输仓储费用赔偿",
    "организатор перевозки": "承运人",
    "хранение": "仓储费",
    "удержания": "扣款",
    "операции на приемке": "入库验收操作",
    "фиксированный коэффициент склада по поставке": "固定仓库供应比例",
    "признак продажи юридическому лицу": "售卖给法人",
    "номер короба для обработки товара": "处理商品箱号",
    "скидка по программе софинансирования": "联合融资折扣",
    "скидка wibes, %": "wibes折扣%",
    "компенсация скидки по программе лояльности": "忠诚度折扣补偿",
    "стоимость участия в программе лояльности": "忠诚度参与费",
    "сумма баллов, удержанных по программе лояльности": "忠诚度扣除积分",
    "id корзины заказа": "订单购物车id",
    "разовое изменение срока перечисления денежных средств": "一次性结算周期变更",
    "id собственной акции продавца с дополнительной скидкой": "卖家额外折扣活动id",
    "размер дополнительной скидки по собственной акции продавца, %": "卖家额外折扣%",
    "способы продажи и тип товара": "销售方式与商品类型",
    "уникальный идентификатор скидки лояльности от продавца": "卖家忠诚度折扣id",
    "размер скидки лояльности от продавца, %": "卖家忠诚度折扣%",
    "id промокода": "优惠码id",
    "скидка за промокод, %": "优惠码折扣%",
    "id подменного артикула": "替换商品id",
    "скидка по подменному артикулу, %": "替换商品折扣%",
    "оптовая скидка для бизнеса, %": "批发折扣%",
}

# Payment type translation
PAYMENT_TYPE_MAP = {
    "продажа": "销售",
    "логистика": "物流",
    "удержание": "广告",
    "хранение": "仓储",
    "штраф": "罚款",
}

CATEGORY_COLUMNS = {
    "销售": ["条形码", "付款依据", "销售日期", "数量", "零售价", "支付给卖家的已售商品金额"],
    "物流": ["条形码", "付款依据", "销售日期", "交付数量", "退货数量", "向买家交付货物的服务"],
    "仓储": ["付款依据", "销售日期", "仓储费"],
    "广告": ["付款依据", "销售日期", "wb物流罚款调整类型", "扣款"],
    "罚款": ["付款依据", "销售日期", "罚款总额", "wb物流罚款调整类型"],
}


def extract_product_code(barcode):
    if not barcode:
        return None
    b = str(barcode).strip().upper()
    m = re.match(r'^([A-Z]+)', b)
    if m:
        return m.group(1)
    return b[:4]


def style_header(ws, row=1):
    hf = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill("solid", fgColor="4472C4")
    thin = Side(style='thin', color='D9D9D9')
    border = Border(bottom=thin)
    for cell in ws[row]:
        cell.font = hf
        cell.fill = hfill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border


def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        cl = col[0].column_letter
        for cell in col:
            val = str(cell.value or "")
            max_len = max(max_len, len(val.encode('utf-8')))
        ws.column_dimensions[cl].width = min(max_len * 0.8 + 2, 35)


# ── Shared processing functions ─────────────────────────

async def _load_purchase_data(db: AsyncSession) -> dict:
    """Load purchase data from the latest purchase file in DB."""
    purchase_data = {}
    purchase_result = await db.execute(
        select(UploadedReport)
        .where(UploadedReport.file_type == "purchase")
        .where(UploadedReport.file_path.isnot(None))
        .order_by(desc(UploadedReport.uploaded_at))
        .limit(1)
    )
    purchase_report = purchase_result.scalar_one_or_none()
    if purchase_report and os.path.exists(purchase_report.file_path):
        try:
            wb_p = load_workbook(purchase_report.file_path, data_only=True)
            ws_p = wb_p.active
            p_headers = []
            name_col = cost_col = head_col = label_col = -1
            for row in ws_p.iter_rows(values_only=True):
                if not row or not any(row):
                    continue
                if not p_headers:
                    p_headers = [str(c or "").strip() for c in row]
                    for i, h in enumerate(p_headers):
                        hl = h.lower()
                        if hl == "品名":
                            name_col = i
                        elif any(kw in hl for kw in ("单套货本", "货本", "成本", "себестоимость")):
                            cost_col = i
                        elif any(kw in hl for kw in ("单套头程", "头程", "运费", "первая миля")):
                            head_col = i
                        elif any(kw in hl for kw in ("单套标签", "标签", "наклейка")):
                            label_col = i
                    continue
                if name_col < 0:
                    continue
                name_val = str(row[name_col] or "").strip() if name_col < len(row) else ""
                if not name_val:
                    continue
                cost_val = float(row[cost_col]) if cost_col >= 0 and cost_col < len(row) and isinstance(row[cost_col], (int, float)) else 0.0
                head_val = float(row[head_col]) if head_col >= 0 and head_col < len(row) and isinstance(row[head_col], (int, float)) else 0.0
                label_val = float(row[label_col]) if label_col >= 0 and label_col < len(row) and isinstance(row[label_col], (int, float)) else 0.0
                purchase_data[name_val] = {"cost": cost_val, "head_freight": head_val, "label_cost": label_val}
        except Exception:
            pass
    return purchase_data


async def _process_single_report(
    file_path: str,
    purchase_data: dict,
    exchange_rate: float,
    tax_rate: float,
    fee_rate: float,
    db: AsyncSession,
    cleanup_ids: list = None,
    cleanup_paths: list = None,
) -> dict:
    """Process a WB report Excel file into multi-sheet output.
    Returns dict with status, filename, download_url, stats.
    """
    try:
        wb_in = load_workbook(file_path, data_only=True)
        raw_sheet = wb_in[wb_in.sheetnames[0]]
        rows_raw = list(raw_sheet.iter_rows(values_only=True))
        if not rows_raw:
            raise HTTPException(status_code=400, detail="工作表为空")

        raw_headers = [str(c or "").strip().lower() for c in rows_raw[0]]
        cn_headers = [HEADER_MAP.get(h, h) for h in raw_headers]
        col_idx = {h: i for i, h in enumerate(cn_headers)}

        wb_out = Workbook()

        payment_col = col_idx.get("付款依据", -1)
        barcode_col = col_idx.get("条形码", -1)
        product_name_col = col_idx.get("商品名称", -1)
        qty_col = col_idx.get("数量", -1)
        for_pay_col = col_idx.get("支付给卖家的已售商品金额", -1)
        retail_col = col_idx.get("零售价", -1)
        delivery_qty_col = col_idx.get("交付数量", -1)
        return_qty_col = col_idx.get("退货数量", -1)
        logistics_col = col_idx.get("向买家交付货物的服务", -1)
        storage_col = col_idx.get("仓储费", -1)
        deduct_col = col_idx.get("扣款", -1)
        penalty_col = col_idx.get("罚款总额", -1)
        adj_type_col = col_idx.get("wb物流罚款调整类型", -1)

        def get_val(row, idx):
            if idx >= 0 and idx < len(row):
                return row[idx]
            return None

        processed_rows = []
        for row in rows_raw[1:]:
            if not row or not any(row):
                continue
            new_row = list(row)
            if payment_col >= 0 and payment_col < len(new_row):
                orig = str(new_row[payment_col] or "").strip().lower()
                translated = PAYMENT_TYPE_MAP.get(orig, new_row[payment_col])
                new_row[payment_col] = translated
            processed_rows.append(new_row)

        ws_proc = wb_out.active
        ws_proc.title = "初处理"
        ws_proc.append(cn_headers)
        for row in processed_rows:
            ws_proc.append(row)
        style_header(ws_proc)
        auto_width(ws_proc)

        # Category sheets
        categorized = {"销售": [], "物流": [], "仓储": [], "广告": [], "罚款": []}
        for row in processed_rows:
            pv = str(get_val(row, payment_col) or "").strip()
            if pv in categorized:
                categorized[pv].append(row)

        CATS = {
            "销售": ["条形码", "付款依据", "销售日期", "数量", "零售价", "支付给卖家的已售商品金额"],
            "物流": ["条形码", "付款依据", "销售日期", "交付数量", "退货数量", "向买家交付货物的服务"],
            "仓储": ["付款依据", "销售日期", "仓储费"],
            "广告": ["付款依据", "销售日期", "wb物流罚款调整类型", "扣款"],
            "罚款": ["付款依据", "销售日期", "罚款总额", "wb物流罚款调整类型"],
        }

        for cat_name, cat_rows in categorized.items():
            if not cat_rows:
                continue
            ws = wb_out.create_sheet(cat_name)
            cols_needed = CATS[cat_name]
            ws.append(cols_needed)
            for row in cat_rows:
                new_row = [get_val(row, col_idx.get(c, -1)) for c in cols_needed]
                ws.append(new_row)
            style_header(ws)
            auto_width(ws)

        # Profit table
        products = {}
        for row in processed_rows:
            pv = str(get_val(row, payment_col) or "").strip()
            barcode = str(get_val(row, barcode_col) or "")
            code = extract_product_code(barcode)
            product_name = str(get_val(row, product_name_col) or "").strip() if product_name_col >= 0 else ""
            if not code:
                continue
            if code not in products:
                products[code] = {"code": code, "name": product_name, "barcode": barcode, "qty": 0, "for_pay": 0.0,
                                  "logistics": 0.0, "delivery_qty": 0, "return_qty": 0}
            p = products[code]
            if product_name and not p["name"]:
                p["name"] = product_name
            if pv == "销售":
                qty = float(get_val(row, qty_col) or 0)
                fp = float(get_val(row, for_pay_col) or 0)
                p["qty"] += qty
                p["for_pay"] += fp
            elif pv == "物流":
                lc = float(get_val(row, logistics_col) or 0)
                p["logistics"] += lc

        total_storage = sum(
            float(get_val(row, storage_col) or 0)
            for row in processed_rows
            if str(get_val(row, payment_col) or "").strip() == "仓储"
        )
        total_qty = sum(p["qty"] for p in products.values()) or 1
        storage_per_unit = total_storage / total_qty

        tax_factor = (1 - tax_rate / 100) * (1 - fee_rate / 100)

        profit_data = []
        for code in sorted(products, key=lambda c: products[c]["for_pay"], reverse=True):
            p = products[code]
            qty = p["qty"]
            for_pay = p["for_pay"]
            logistics = p["logistics"]

            avg_price = round(for_pay / qty, 2) if qty > 0 else 0
            avg_log = round(logistics / qty, 2) if qty > 0 else 0
            storage_fee = round(qty * storage_per_unit, 2)

            purch = purchase_data.get(code, None) or purchase_data.get(p["name"], None) or {"cost": 0, "head_freight": 0, "label_cost": 0}
            cost_per_unit = purch["cost"]
            head_per_unit = purch["head_freight"]
            label_cost_rub = purch.get("label_cost", 0)  # 卢布
            cost_total = round(cost_per_unit * qty, 2)
            head_total = round(head_per_unit * qty, 2)
            label_total_rub = round(label_cost_rub * qty, 2)

            total_sum = round(for_pay - logistics - storage_fee - label_total_rub, 2)
            after_tax = round(total_sum * tax_factor, 2)
            to_cny = round(after_tax / exchange_rate, 2)

            total_profit = round(after_tax - cost_total - head_total, 2)
            profit_data.append([code, int(qty), avg_price, round(for_pay, 2),
                                0, avg_log, round(logistics, 2), storage_fee,
                                cost_per_unit, head_per_unit, label_cost_rub, cost_total, head_total, label_total_rub,
                                total_sum, after_tax, to_cny,
                                total_profit,
                                round(total_profit / qty, 2) if qty > 0 else 0])

        ws_profit = wb_out.create_sheet("利润表")
        profit_h = ["品名", "数量", "平均单套售价", "支付金额",
                    "退货金额", "平均单套物流", "物流费", "仓储费",
                    "单套货本", "单套头程", "单套标签(₽)", "货本总计", "头程总计", "标签总计(₽)",
                    "总和", "扣税和手续费后", "汇率转人民币", "总利润", "单个利润"]
        ws_profit.append(profit_h)
        for row in profit_data:
            ws_profit.append(row)
        style_header(ws_profit)
        auto_width(ws_profit)

        fname = f"WB_processed_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        out_path = os.path.join(UPLOAD_DIR, fname)
        wb_out.save(out_path)

        # Cleanup source files
        if cleanup_paths:
            for p in cleanup_paths:
                try:
                    if os.path.exists(p):
                        os.remove(p)
                except Exception:
                    pass
        if cleanup_ids:
            for rid in cleanup_ids:
                try:
                    r = await db.execute(select(UploadedReport).where(UploadedReport.id == rid))
                    rec = r.scalar_one_or_none()
                    if rec:
                        await db.delete(rec)
                except Exception:
                    pass
        # Remove old processed files
        for old_f in os.listdir(UPLOAD_DIR):
            if old_f != fname and old_f.endswith(".xlsx"):
                try:
                    os.remove(os.path.join(UPLOAD_DIR, old_f))
                except Exception:
                    pass
        await db.commit()

        return {
            "status": "success",
            "filename": fname,
            "download_url": f"/api/v1/reports/process/download/{fname}",
            "stats": {
                "total_rows": len(processed_rows),
                "sales_rows": len(categorized.get("销售", [])),
                "logistics_rows": len(categorized.get("物流", [])),
                "storage_rows": len(categorized.get("仓储", [])),
                "ad_rows": len(categorized.get("广告", [])),
                "penalty_rows": len(categorized.get("罚款", [])),
                "products": len(profit_data),
                "total_storage": round(total_storage, 2),
                "exchange_rate": exchange_rate,
                "tax_rate": tax_rate,
                "fee_rate": fee_rate,
            }
        }

    except Exception as e:
        import traceback
        raise HTTPException(status_code=400, detail=f"处理失败: {str(e)}\n{traceback.format_exc()}")


@router.post("/reports/process")
async def process_reports(
    exchange_rate: float = 12.5,
    tax_rate: float = 12,
    fee_rate: float = 7,
    db: AsyncSession = Depends(get_db),
):
    """Process the latest platform report → multi-sheet Excel download.
    Uses latest purchase file for cost/head_freight if available.
    Accepts: exchange_rate, tax_rate(%), fee_rate(%).
    """
    # Find latest platform report
    result = await db.execute(
        select(UploadedReport)
        .where(UploadedReport.file_type == "platform")
        .where(UploadedReport.file_path.isnot(None))
        .order_by(desc(UploadedReport.uploaded_at))
        .limit(1)
    )
    report = result.scalar_one_or_none()
    if not report:
        raise HTTPException(status_code=404, detail="没有找到已上传的平台报表")
    
    file_path = report.file_path
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail=f"平台文件不存在: {file_path}")

    # Find latest purchase file — 3 columns: 品名, 单套货本, 单套头程
    purchase_data = await _load_purchase_data(db)
    
    return await _process_single_report(file_path, purchase_data, exchange_rate, tax_rate, fee_rate, db, cleanup_ids=[report.id], cleanup_paths=[file_path])


@router.get("/reports/process/download/{filename}")
async def download_processed_file(filename: str):
    """Download a processed report file."""
    file_path = os.path.join(UPLOAD_DIR, filename)
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="文件不存在或已过期")
    return FileResponse(
        file_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=filename
    )
